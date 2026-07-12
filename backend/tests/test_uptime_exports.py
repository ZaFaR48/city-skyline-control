from __future__ import annotations

from datetime import datetime, timedelta, timezone
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from starlette.requests import Request

from app.models import OperationalRegion, Station, StationStatusEvent, User
from app.routers.reports import export_uptime_csv, export_uptime_xlsx
from app.services.uptime_reports import build_uptime_report


def request() -> Request:
    return Request({"type": "http", "method": "GET", "path": "/api/reports/uptime", "headers": [], "client": ("127.0.0.1", 1)})


async def payload(response) -> bytes:
    return b"".join([chunk async for chunk in response.body_iterator])


async def report_station(db, code: str) -> Station:
    city = (await db.execute(select(OperationalRegion).where(OperationalRegion.code == "dushanbe"))).scalar_one()
    row = Station(station_code=code, name=f"Report {code}", city_id=city.id, address="", approved_at=datetime.now(timezone.utc), is_active=True, is_archived=False)
    db.add(row)
    await db.flush()
    return row


@pytest.mark.asyncio
async def test_report_intervals_sum_to_range_and_unknown_is_not_offline(db):
    end = datetime.now(timezone.utc) - timedelta(minutes=1)
    start = end - timedelta(hours=4)
    row = await report_station(db, "96001")
    db.add_all([
        StationStatusEvent(station_id=row.id, previous_status="unknown", new_status="online", source="ping", started_at=start + timedelta(hours=1), ended_at=start + timedelta(hours=3)),
        StationStatusEvent(station_id=row.id, previous_status="online", new_status="offline", source="ping", started_at=start + timedelta(hours=3), ended_at=end),
    ])
    await db.flush()
    result = (await build_uptime_report(db, start=start, end=end, station_id=row.id))[0]
    assert result.online_seconds + result.degraded_seconds + result.offline_seconds + result.unknown_seconds == result.total_monitored_seconds
    assert result.unknown_seconds == 3600
    assert result.offline_seconds == 3600
    assert result.availability_percentage == pytest.approx(66.67, abs=0.01)
    assert result.data_coverage_percentage == 75.0
    assert result.outages == 1 and result.longest_outage_seconds == 3600


@pytest.mark.asyncio
async def test_csv_and_xlsx_exports_have_attachment_headers_and_valid_content(db):
    end = datetime.now(timezone.utc) - timedelta(minutes=1)
    start = end - timedelta(hours=1)
    actor = User(username="export-admin", email="export@test.invalid", hashed_password="x", role="admin", is_active=True)
    db.add(actor)
    await db.flush()
    csv_response = await export_uptime_csv(start, end, request(), db=db, user=actor)
    csv_bytes = await payload(csv_response)
    assert csv_response.media_type.startswith("text/csv")
    assert "attachment" in csv_response.headers["content-disposition"]
    assert csv_bytes.startswith(b"\xef\xbb\xbf") and b"station_code" in csv_bytes

    xlsx_response = await export_uptime_xlsx(start, end, request(), db=db, user=actor)
    xlsx_bytes = await payload(xlsx_response)
    assert xlsx_response.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment" in xlsx_response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(xlsx_bytes), read_only=True)
    assert workbook["Uptime"].cell(1, 2).value == "station_code"
